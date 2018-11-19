from openpyxl import load_workbook

fileDir = "c:/Users/e1ze4m0/Documents/Indivior Program/Indivior Hub/SP Status/SP PA Check In 10.09.xlsx"

wb = load_workbook(filename=fileDir)

print(wb.sheetnames)
#print(wb.get_sheet_names())

#for sheet in wb.worksheets:
#    print(*sheet, sep="\n")
